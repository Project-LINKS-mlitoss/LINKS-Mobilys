"""
Core services for Ridership processing
Handles Excel file upload, parsing, validation, and database storage
"""

import uuid
import logging
from datetime import datetime, time, date
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple
from zoneinfo import ZoneInfo

from django.utils import timezone
from rest_framework import status

# JST timezone for ridership datetime processing
JST = ZoneInfo("Asia/Tokyo")

from gtfs.services.base import transactional


import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from dateutil import parser as dateutil_parser

# Import models - adjust the import path according to your project structure
from gtfs.models import (
    Scenario,
    RidershipUpload,
    RidershipUploadError,
    RidershipRecord,
    Trips,
    StopTimes,
    Stops,
    Routes,
    Calendar,
    CalendarDates,
)

# Import serializers (request/response kept separate)
from gtfs.serializers.request.ridership_request import RidershipUploadRequestSerializer
from gtfs.serializers.response.ridership_response import (
    RidershipUploadDetailSerializer,
    RidershipUploadListSerializer,
    RidershipUploadResponseSerializer,
)

logger = logging.getLogger(__name__)


class _GtfsDataEnricherImpl:
    """
    Enriches ridership records with GTFS data by matching trips based on
    boarding/alighting stops and payment time.
    
    Algorithm (based on IC card data matching specification):
    1. Extract valid service_ids for the travel date from calendar and calendar_dates
    2. Find trips for boarding station (by code or name) with time window matching
    3. Find trips for alighting station (by code or name) with time window matching  
    4. Match trip_id that appears in both boarding and alighting candidates
    5. Validate direction (boarding sequence < alighting sequence)
    6. Select best match based on time difference
    """
    
    # Default time window settings (in minutes) - used when max_tolerance_time is not set
    INITIAL_TIME_WINDOW = 5  # ±5 minutes
    EXPANDED_TIME_WINDOWS = [10, 15, 20]  # Fallback windows
    
    def __init__(self, scenario: Scenario, max_tolerance_time: Optional[int] = None):
        """
        Initialize the enricher.
        
        Args:
            scenario: The scenario to search GTFS data from
            max_tolerance_time: Optional maximum time tolerance in minutes.
                               If provided, search directly with this window (no expanding).
                               If None, use expanding window logic (5 -> 10 -> 15 -> 20).
        """
        self.scenario = scenario
        self.max_tolerance_time = max_tolerance_time
        
        # Cache for stop names -> stop_ids mapping
        self._stop_name_to_ids: Dict[str, List[str]] = {}
        # Cache for stop_id -> stop_name
        self._stop_id_to_name: Dict[str, str] = {}
        # Cache for route info
        self._route_cache: Dict[str, Dict] = {}
        # Cache for trip -> route mapping
        self._trip_route_cache: Dict[str, str] = {}
        # Cache for service_ids by date
        self._service_id_cache: Dict[str, set] = {}
        
        # Preload caches
        self._preload_stops()
        self._preload_routes()
        self._preload_trip_routes()
    
    def _preload_stops(self):
        """Preload all stops for the scenario into cache"""
        stops = Stops.objects.filter(scenario=self.scenario).values('stop_id', 'stop_name')
        for stop in stops:
            stop_id = stop['stop_id']
            stop_name = stop['stop_name'] or ''
            
            # stop_id -> stop_name
            self._stop_id_to_name[stop_id] = stop_name
            
            # stop_name -> list of stop_ids (multiple stops can have same name)
            if stop_name:
                normalized_name = stop_name.strip()
                if normalized_name not in self._stop_name_to_ids:
                    self._stop_name_to_ids[normalized_name] = []
                self._stop_name_to_ids[normalized_name].append(stop_id)
    
    def _preload_routes(self):
        """Preload all routes for the scenario into cache"""
        routes = Routes.objects.filter(scenario=self.scenario).values(
            'route_id', 'route_short_name', 'route_long_name'
        )
        for route in routes:
            self._route_cache[route['route_id']] = {
                'route_id': route['route_id'],
                'route_short_name': route['route_short_name'] or '',
                'route_long_name': route['route_long_name'] or '',
                'route_name': route['route_long_name'] or route['route_short_name'] or '',
            }
    
    def _preload_trip_routes(self):
        """Preload trip -> route mapping"""
        trips = Trips.objects.filter(scenario=self.scenario).values('trip_id', 'route_id')
        for trip in trips:
            self._trip_route_cache[trip['trip_id']] = trip['route_id']
    
    def get_stop_name(self, stop_id: str) -> str:
        """Get stop name from cache"""
        return self._stop_id_to_name.get(stop_id, '')
    
    def get_stop_ids_by_name(self, stop_name: str) -> List[str]:
        """Get list of stop_ids that match the given stop name"""
        if not stop_name:
            return []
        normalized_name = stop_name.strip()
        return self._stop_name_to_ids.get(normalized_name, [])
    
    def get_route_info(self, route_id: str) -> Dict:
        """Get route info from cache"""
        return self._route_cache.get(route_id, {})
    
    def get_route_id_for_trip(self, trip_id: str) -> str:
        """Get route_id for a trip from cache"""
        return self._trip_route_cache.get(trip_id, '')
    
    def _get_valid_service_ids(self, travel_date: date) -> set:
        """
        Step 1: Get valid service_ids for the travel date.
        
        Logic:
        1. Check calendar.txt for services active on the day of week within date range
        2. Apply calendar_dates.txt exceptions (add/remove services)
        
        Args:
            travel_date: The date to check service availability
            
        Returns:
            Set of valid service_ids for the given date
        """
        # Check cache first
        date_str = travel_date.isoformat()
        if date_str in self._service_id_cache:
            return self._service_id_cache[date_str]
        
        valid_service_ids = set()
        
        # Get day of week (0=Monday, 6=Sunday)
        day_of_week = travel_date.weekday()
        day_field_map = {
            0: 'monday',
            1: 'tuesday',
            2: 'wednesday',
            3: 'thursday',
            4: 'friday',
            5: 'saturday',
            6: 'sunday',
        }
        day_field = day_field_map[day_of_week]
        
        # Step 1: Get services from calendar.txt that are active on this day
        calendar_filter = {
            'scenario': self.scenario,
            'start_date__lte': travel_date,
            'end_date__gte': travel_date,
            day_field: True,  # e.g., monday=True for Monday
        }
        
        try:
            calendar_services = Calendar.objects.filter(**calendar_filter).values_list('service_id', flat=True)
            valid_service_ids.update(calendar_services)
        except Exception as e:
            logger.warning(f"Error querying calendar: {e}")
        
        # Step 2: Apply calendar_dates.txt exceptions
        try:
            exceptions = CalendarDates.objects.filter(
                scenario=self.scenario,
                date=travel_date
            ).values('service_id', 'exception_type')
            
            for exception in exceptions:
                service_id = exception['service_id']
                exception_type = exception['exception_type']
                
                if exception_type == 1:  # Service added
                    valid_service_ids.add(service_id)
                elif exception_type == 2:  # Service removed
                    valid_service_ids.discard(service_id)
        except Exception as e:
            logger.warning(f"Error querying calendar_dates: {e}")
        
        # Cache the result
        self._service_id_cache[date_str] = valid_service_ids
        
        return valid_service_ids
    
    def _get_valid_trip_ids(self, travel_date: date) -> set:
        """
        Get valid trip_ids for the travel date based on service availability.
        
        Args:
            travel_date: The date to check trip availability
            
        Returns:
            Set of valid trip_ids for the given date
        """
        valid_service_ids = self._get_valid_service_ids(travel_date)
        
        if not valid_service_ids:
            # If no service_ids found, return empty set (trip not found)
            logger.debug(f"No valid service_ids for date {travel_date}, skipping enrichment")
            return set()
        
        valid_trips = Trips.objects.filter(
            scenario=self.scenario,
            service_id__in=valid_service_ids
        ).values_list('trip_id', flat=True)
        
        return set(valid_trips)
    
    def _time_to_seconds(self, t) -> int:
        """Convert time object to seconds from midnight"""
        if t is None:
            return 0
        if hasattr(t, 'hour'):
            return t.hour * 3600 + t.minute * 60 + t.second
        return 0
    
    def _seconds_to_time(self, seconds: int) -> time:
        """Convert seconds from midnight to time object"""
        hours = (seconds // 3600) % 24
        minutes = (seconds % 3600) // 60
        secs = seconds % 60
        return time(hours, minutes, secs)
    
    def _seconds_to_time_str(self, seconds: int) -> str:
        """Convert seconds from midnight to HH:MM:SS string"""
        hours = seconds // 3600
        minutes = (seconds % 3600) // 60
        secs = seconds % 60
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"
    
    def _find_stop_times_by_station(
        self,
        stop_code: str,
        stop_name: str,
        target_time: datetime,
        valid_trip_ids: set,
        time_field: str = 'arrival_time',
        time_window_minutes: int = 5
    ) -> List[Dict]:
        """
        Find stop_times records matching station (by code or name) within time window.
        
        Args:
            stop_code: Station code (stop_id)
            stop_name: Station name (used if code is empty)
            target_time: Target datetime to match
            valid_trip_ids: Set of valid trip_ids based on service date
            time_field: Which time field to use ('arrival_time' or 'departure_time')
            time_window_minutes: Time window in minutes (±)
            
        Returns:
            List of matching stop_times with trip info
        """
        if not valid_trip_ids:
            return []
        
        # Determine which stop_ids to search
        stop_ids_to_search = []
        
        if stop_code:
            # Primary: search by stop_code
            stop_ids_to_search = [stop_code]
        elif stop_name:
            # Fallback: search by stop_name if no code provided
            stop_ids_to_search = self.get_stop_ids_by_name(stop_name)
        
        if not stop_ids_to_search:
            return []
        
        # Calculate time window
        target_time_only = target_time.time() if hasattr(target_time, 'time') else target_time
        target_seconds = self._time_to_seconds(target_time_only)
        window_seconds = time_window_minutes * 60
        
        min_seconds = target_seconds - window_seconds
        max_seconds = target_seconds + window_seconds
        
        # Handle negative time (previous day) - just use 0
        if min_seconds < 0:
            min_seconds = 0
        
        min_time = self._seconds_to_time(min_seconds)
        max_time = self._seconds_to_time(max_seconds % 86400)  # Wrap around midnight
        
        # Build query
        base_filter = {
            'scenario': self.scenario,
            'stop_id__in': stop_ids_to_search,
            'trip_id__in': valid_trip_ids,
        }
        
        # Handle time window query
        if min_seconds < max_seconds % 86400:
            # Normal case: min < max
            base_filter[f'{time_field}__gte'] = min_time
            base_filter[f'{time_field}__lte'] = max_time
        else:
            # Wrap around midnight case - query in two parts
            # This is simplified; for now just use the range
            base_filter[f'{time_field}__gte'] = min_time
        
        stop_times = StopTimes.objects.filter(**base_filter).values(
            'trip_id', 'stop_id', 'stop_sequence', 
            'arrival_time', 'departure_time'
        )
        
        results = []
        for st in stop_times:
            time_value = st.get(time_field)
            if time_value:
                time_seconds = self._time_to_seconds(time_value)
                time_diff = abs(time_seconds - target_seconds)
                results.append({
                    'trip_id': st['trip_id'],
                    'stop_id': st['stop_id'],
                    'stop_sequence': st['stop_sequence'],
                    'arrival_time': st['arrival_time'],
                    'departure_time': st['departure_time'],
                    'time_diff_seconds': time_diff,
                })
        
        return results
    
    def _find_boarding_candidates(
        self,
        stop_code: str,
        stop_name: str,
        boarding_time: datetime,
        valid_trip_ids: set
    ) -> List[Dict]:
        """
        Step 2: Find trip candidates for boarding station.
        
        Uses departure_time because passenger boards when vehicle departs from stop.
        Uses either fixed time window (if max_tolerance_time set) or expanding window.
        """
        if self.max_tolerance_time is not None:
            # Use fixed time window directly
            return self._find_stop_times_by_station(
                stop_code=stop_code,
                stop_name=stop_name,
                target_time=boarding_time,
                valid_trip_ids=valid_trip_ids,
                time_field='departure_time',  # FIXED: Use departure_time for boarding
                time_window_minutes=self.max_tolerance_time
            )
        
        # Use expanding time window logic
        candidates = self._find_stop_times_by_station(
            stop_code=stop_code,
            stop_name=stop_name,
            target_time=boarding_time,
            valid_trip_ids=valid_trip_ids,
            time_field='departure_time',  # FIXED: Use departure_time for boarding
            time_window_minutes=self.INITIAL_TIME_WINDOW
        )
        
        if candidates:
            return candidates
        
        # Try expanded time windows
        for window in self.EXPANDED_TIME_WINDOWS:
            candidates = self._find_stop_times_by_station(
                stop_code=stop_code,
                stop_name=stop_name,
                target_time=boarding_time,
                valid_trip_ids=valid_trip_ids,
                time_field='departure_time',  # FIXED: Use departure_time for boarding
                time_window_minutes=window
            )
            if candidates:
                return candidates
        
        return []
    
    def _find_alighting_candidates(
        self,
        stop_code: str,
        stop_name: str,
        alighting_time: datetime,
        valid_trip_ids: set
    ) -> List[Dict]:
        """
        Step 3: Find trip candidates for alighting station.
        
        Uses arrival_time because passenger alights when vehicle arrives at stop.
        Uses either fixed time window (if max_tolerance_time set) or expanding window.
        """
        if self.max_tolerance_time is not None:
            # Use fixed time window directly
            return self._find_stop_times_by_station(
                stop_code=stop_code,
                stop_name=stop_name,
                target_time=alighting_time,
                valid_trip_ids=valid_trip_ids,
                time_field='arrival_time',  # FIXED: Use arrival_time for alighting
                time_window_minutes=self.max_tolerance_time
            )
        
        # Use expanding time window logic
        candidates = self._find_stop_times_by_station(
            stop_code=stop_code,
            stop_name=stop_name,
            target_time=alighting_time,
            valid_trip_ids=valid_trip_ids,
            time_field='arrival_time',  # FIXED: Use arrival_time for alighting
            time_window_minutes=self.INITIAL_TIME_WINDOW
        )
        
        if candidates:
            return candidates
        
        # Try expanded time windows
        for window in self.EXPANDED_TIME_WINDOWS:
            candidates = self._find_stop_times_by_station(
                stop_code=stop_code,
                stop_name=stop_name,
                target_time=alighting_time,
                valid_trip_ids=valid_trip_ids,
                time_field='arrival_time',  # FIXED: Use arrival_time for alighting
                time_window_minutes=window
            )
            if candidates:
                return candidates
        
        return []
    
    def _select_best_stop_sequence(self, candidates: List[Dict], target_time: datetime) -> Optional[Dict]:
        """
        For circular routes where same stop appears multiple times,
        select the stop_sequence with closest time to target.
        """
        if not candidates:
            return None
        
        # Sort by time difference and return the closest
        sorted_candidates = sorted(candidates, key=lambda x: x.get('time_diff_seconds', float('inf')))
        return sorted_candidates[0]
    
    def find_matching_trip(
        self,
        boarding_station_code: str,
        boarding_station_name: str,
        alighting_station_code: str,
        alighting_station_name: str,
        boarding_time: datetime,
        alighting_time: datetime = None,
        payment_at: datetime = None
    ) -> Optional[Dict]:
        """
        Find the best matching trip based on the IC card matching algorithm.
        
        Algorithm:
        1. Get valid service_ids for the travel date from calendar and calendar_dates
        2. Get valid trip_ids based on service_ids
        3. Find boarding candidates (trips passing through boarding station at ~boarding_time)
        4. Find alighting candidates (trips passing through alighting station at ~alighting_time)
        5. Match trip_id that appears in both lists
        6. Validate direction (boarding sequence < alighting sequence)
        7. Select best match based on time difference
        
        Returns:
            Dict with trip_id, route_id, route_name, boarding_at, alighting_at,
            boarding_stop_sequence, alighting_stop_sequence
            or None if no matching trip found
        """
        # Determine travel date from boarding_time (primary) or alighting_time
        travel_date = None
        if boarding_time and isinstance(boarding_time, datetime):
            travel_date = boarding_time.date()
        elif alighting_time and isinstance(alighting_time, datetime):
            travel_date = alighting_time.date()
        
        if not travel_date:
            logger.debug("No travel_date available for trip matching")
            return None
        
        # Use alighting_time for alighting search, boarding_time for boarding search
        effective_alighting_time = alighting_time
        effective_boarding_time = boarding_time
        
        if not effective_boarding_time or not effective_alighting_time:
            return None
        
        # Step 1 & 2: Get valid trip_ids based on calendar/service date
        valid_trip_ids = self._get_valid_trip_ids(travel_date)
        
        if not valid_trip_ids:
            logger.debug(f"No valid trips for date {travel_date}")
            return None
        
        logger.debug(f"Found {len(valid_trip_ids)} valid trips for date {travel_date}")
        
        # Step 3: Find alighting candidates first (using alighting_time)
        alighting_candidates = self._find_alighting_candidates(
            stop_code=alighting_station_code,
            stop_name=alighting_station_name,
            alighting_time=effective_alighting_time,
            valid_trip_ids=valid_trip_ids
        )
        
        if not alighting_candidates:
            logger.debug("No alighting candidates found")
            return None
        
        # Get trip_ids from alighting candidates
        alighting_trip_ids = {c['trip_id'] for c in alighting_candidates}
        
        # Step 4: Find boarding candidates (filtered by alighting trip_ids)
        boarding_candidates = self._find_boarding_candidates(
            stop_code=boarding_station_code,
            stop_name=boarding_station_name,
            boarding_time=effective_boarding_time,
            valid_trip_ids=alighting_trip_ids  # Only search trips that reach alighting station
        )
        
        if not boarding_candidates:
            logger.debug("No boarding candidates found")
            return None
        
        # Step 5: Find trip_ids that appear in both boarding and alighting candidates
        boarding_trip_ids = {c['trip_id'] for c in boarding_candidates}
        matching_trip_ids = boarding_trip_ids & alighting_trip_ids
        
        if not matching_trip_ids:
            logger.debug("No matching trip_ids found in both boarding and alighting")
            return None
        
        # Step 6: For each matching trip_id, calculate total time difference
        # and ensure boarding comes before alighting (stop_sequence check)
        valid_matches = []
        
        # Group candidates by trip_id
        boarding_by_trip = {}
        for candidate in boarding_candidates:
            trip_id = candidate['trip_id']
            if trip_id not in boarding_by_trip:
                boarding_by_trip[trip_id] = []
            boarding_by_trip[trip_id].append(candidate)
        
        alighting_by_trip = {}
        for candidate in alighting_candidates:
            trip_id = candidate['trip_id']
            if trip_id not in alighting_by_trip:
                alighting_by_trip[trip_id] = []
            alighting_by_trip[trip_id].append(candidate)
        
        for trip_id in matching_trip_ids:
            # For circular routes, find best combination of boarding/alighting
            for boarding in boarding_by_trip.get(trip_id, []):
                for alighting in alighting_by_trip.get(trip_id, []):
                    # Validate direction: boarding must come before alighting
                    if boarding['stop_sequence'] < alighting['stop_sequence']:
                        total_diff = boarding.get('time_diff_seconds', 0) + alighting.get('time_diff_seconds', 0)
                        valid_matches.append({
                            'trip_id': trip_id,
                            'boarding_info': boarding,
                            'alighting_info': alighting,
                            'total_time_diff': total_diff,
                        })
        
        if not valid_matches:
            logger.debug("No valid matches with correct direction (boarding_sequence < alighting_sequence)")
            return None
        
        # Step 7: Select best match (lowest total time difference)
        best_match = min(valid_matches, key=lambda x: x['total_time_diff'])
        
        trip_id = best_match['trip_id']
        boarding_info = best_match['boarding_info']
        alighting_info = best_match['alighting_info']
        
        # Get route info
        route_id = self.get_route_id_for_trip(trip_id)
        route_info = self.get_route_info(route_id)
        
        return {
            'trip_id': trip_id,
            'route_id': route_id,
            'route_name': route_info.get('route_name', ''),
            'boarding_at': boarding_info['departure_time'],
            'alighting_at': alighting_info['arrival_time'],
            'boarding_stop_sequence': boarding_info['stop_sequence'],
            'alighting_stop_sequence': alighting_info['stop_sequence'],
            'boarding_stop_id': boarding_info['stop_id'],
            'alighting_stop_id': alighting_info['stop_id'],
        }
    
    def enrich_record_data(
        self,
        record_data: Dict,
        boarding_station_code: str,
        alighting_station_code: str,
        payment_at: datetime
    ) -> Tuple[Dict, bool]:
        """
        Enrich record data with GTFS data if fields are empty.

        Fields to potentially fill:
        - trip_code (from trip_id)
        - route_id
        - route_name
        - boarding_station_name (from stop_name)
        - boarding_station_code (if matched by name)
        - boarding_station_sequence
        - alighting_station_name (from stop_name)
        - alighting_station_code (if matched by name)
        - alighting_station_sequence

        Returns:
            Tuple[Dict, bool]: (enriched_record_data, trip_found)
            - trip_found: True if trip matching was successful, False if needed but not found
        """
        # Get station names from record data (may be provided even if code is not)
        boarding_station_name = record_data.get('boarding_station_name', '')
        alighting_station_name = record_data.get('alighting_station_name', '')
        
        # Fill station names from stops table if code is provided but name is empty
        if boarding_station_code and not boarding_station_name:
            record_data['boarding_station_name'] = self.get_stop_name(boarding_station_code)
            boarding_station_name = record_data['boarding_station_name']
        
        if alighting_station_code and not alighting_station_name:
            record_data['alighting_station_name'] = self.get_stop_name(alighting_station_code)
            alighting_station_name = record_data['alighting_station_name']
        
        # Get boarding/alighting times from record data
        boarding_time = record_data.get('boarding_at')
        alighting_time = record_data.get('alighting_at')

        # Check if we need to find trip info
        needs_trip_info = (
            not record_data.get('trip_code') or
            not record_data.get('route_id') or
            not record_data.get('route_name')
        )

        # Check if we have enough info to search (either code or name for both stations)
        has_boarding_info = bool(boarding_station_code or boarding_station_name)
        has_alighting_info = bool(alighting_station_code or alighting_station_name)
        has_time_info = bool(boarding_time and alighting_time)

        # Track if trip matching was attempted and failed
        trip_found = True  # Default to True (no issue)

        if needs_trip_info and has_boarding_info and has_alighting_info and has_time_info:
            trip_info = self.find_matching_trip(
                boarding_station_code=boarding_station_code,
                boarding_station_name=boarding_station_name,
                alighting_station_code=alighting_station_code,
                alighting_station_name=alighting_station_name,
                boarding_time=boarding_time,
                alighting_time=alighting_time,
                payment_at=payment_at
            )

            if trip_info:
                # Fill in missing fields from trip info
                if not record_data.get('trip_code'):
                    record_data['trip_code'] = trip_info['trip_id']

                if not record_data.get('route_id'):
                    record_data['route_id'] = trip_info['route_id']

                if not record_data.get('route_name'):
                    record_data['route_name'] = trip_info['route_name']

                # Fill boarding station code if it was matched by name
                if not boarding_station_code and trip_info.get('boarding_stop_id'):
                    record_data['boarding_station_code'] = trip_info['boarding_stop_id']

                # Fill alighting station code if it was matched by name
                if not alighting_station_code and trip_info.get('alighting_stop_id'):
                    record_data['alighting_station_code'] = trip_info['alighting_stop_id']

                # Fill stop sequences
                if trip_info.get('boarding_stop_sequence'):
                    record_data['boarding_station_sequence'] = trip_info['boarding_stop_sequence']

                if trip_info.get('alighting_stop_sequence'):
                    record_data['alighting_station_sequence'] = trip_info['alighting_stop_sequence']
            else:
                # Trip matching was attempted but failed
                trip_found = False

        return record_data, trip_found
    

class _RidershipRecordProcessorImpl:
    """
    Processor class for handling ridership record parsing and validation
    """
    
    # Validation mode choices
    VALIDATION_MODE_RAILWAY = 'railway'
    VALIDATION_MODE_BUS_IC = 'bus_ic'
    VALIDATION_MODE_BUS_CASH = 'bus_cash'
    
    VALIDATION_MODE_CHOICES = [
        VALIDATION_MODE_RAILWAY,
        VALIDATION_MODE_BUS_IC,
        VALIDATION_MODE_BUS_CASH,
    ]
    
    # Excel column name to model field name mapping
    COLUMN_MAPPING = {
        'ic_card_agency_identification_code': 'ic_card_agency_identification_code',
        'ic_card_issuer_code': 'ic_card_issuer_code',
        'ic_card_issuer_name': 'ic_card_issuer_name',
        'ic_card_feature_type': 'ic_card_feature_type',
        'ticket_type_area_code': 'ticket_type_area_code',
        'ticket_type': 'ticket_type',
        'ticket_type_name': 'ticket_type_name',
        'ticket_valid_start_date': 'ticket_valid_start_date',
        'ticket_valid_end_date': 'ticket_valid_end_date',
        'ridership_record_id': 'ridership_record_id',
        'transportation_mode_code': 'transportation_mode_code',
        'ic_card_usage_detail_id': 'ic_card_usage_detail_id',
        'operating_agency_code': 'operating_agency_code',
        'operating_agency_name': 'operating_agency_name',
        'serviced_office_code': 'serviced_office_code',
        'serviced_office_name': 'serviced_office_name',
        'route_pattern_id': 'route_pattern_id',
        'route_pattern_number': 'route_pattern_number',
        'service_line_name': 'service_line_name',
        'route_id': 'route_id',
        'route_name': 'route_name',
        'trip_code': 'trip_code',
        'timetable_number': 'timetable_number',
        'vihicle_number': 'vehicle_number',  # Handle typo in Excel
        'vehicle_number': 'vehicle_number',
        'operation_type': 'operation_type',
        'operation_detail_type': 'operation_detail_type',
        'boarding_area_code': 'boarding_area_code',
        'boarding_station_sequence': 'boarding_station_sequence',
        'boarding_station_code': 'boarding_station_code',
        'boarding_station_name': 'boarding_station_name',
        'boarding_at': 'boarding_at',
        'transfer_area_code_list': 'transfer_area_code_list',
        'transfer_station_code_list': 'transfer_station_code_list',
        'alighting_area_code': 'alighting_area_code',
        'alighting_station_sequence': 'alighting_station_sequence',
        'alighting_station_code': 'alighting_station_code',
        'alighting_station_name': 'alighting_station_name',
        'alighting_at': 'alighting_at',
        'payment_at': 'payment_at',
        'adult_challenged_passenger_count': 'adult_challenged_passenger_count',
        'adult_passenger_count': 'adult_passenger_count',
        'child_challenged_passenger_count': 'child_challenged_passenger_count',
        'child_passenger_count': 'child_passenger_count',
        'passenger_classification_type': 'passenger_classification_type',
    }
    
    # Required fields - time fields are always required
    # Station fields now use conditional logic (code OR name)
    REQUIRED_TIME_FIELDS = [
        'boarding_at',
        'alighting_at',
    ]
    
    # Conditional required fields - code OR name must be present
    # These are validated separately with fallback logic
    CONDITIONAL_STATION_FIELDS = [
        # (code_field, name_field, display_name)
        ('boarding_station_code', 'boarding_station_name', '乗車駅(停留所)'),
        ('alighting_station_code', 'alighting_station_name', '降車駅(停留所)'),
    ]
    
    # Required fields based on validation mode (same for all modes now)
    # Only time fields are strictly required - station uses fallback
    REQUIRED_FIELDS_BY_MODE = {
        VALIDATION_MODE_RAILWAY: REQUIRED_TIME_FIELDS,
        VALIDATION_MODE_BUS_IC: REQUIRED_TIME_FIELDS,
        VALIDATION_MODE_BUS_CASH: REQUIRED_TIME_FIELDS,
    }
    
    # Human-readable field names for error messages (Japanese)
    FIELD_DISPLAY_NAMES = {
        'ic_card_agency_identification_code': 'ICカード識別コード',
        'ridership_record_id': '乗降実績ID',
        'boarding_station_code': '乗車駅(停留所)コード',
        'boarding_station_name': '乗車駅(停留所)名',
        'alighting_station_code': '降車駅(停留所)コード',
        'alighting_station_name': '降車駅(停留所)名',
        'boarding_at': '乗車日時',
        'alighting_at': '降車日時',
        'payment_at': '精算日時',
        'serviced_office_code': '営業所コード',
        'route_pattern_number': '系統番号',
        'adult_passenger_count': '大人利用者数',
        'child_passenger_count': '小児利用者数',
    }
    
    # Enum field definitions with valid values
    ENUM_FIELDS = {
        'ic_card_feature_type': [
            'SF_CREDIT_AUTOCHARGE',
            'SF_ONLY',
            'SF_NO_CREDIT',
        ],
        'ticket_type': [
            'SINGLE_TRIP_TICKET',
            'STUDENT_PASS',
            'COMMUTER_PASS',
            'EXCURSION_TICKET',
            'ONE_DAY_PASS',
            'SHORT_TERM_PASS',
        ],
        'operation_type': [
            'ENTRY',
            'EXIT',
        ],
        'operation_detail_type': [
            'DISCOUNT_TRANSFER',
            'SPECIFIC_FARE',
            'ADDITIONAL_FARE',
        ],
        'passenger_classification_type': [
            'ADULT',
            'CHILD',
            'INFANT',
            'SENIOR',
            'DISABLED',
        ],
    }
    
    # Integer fields
    INTEGER_FIELDS = [
        'ridership_record_id',
        'ic_card_usage_detail_id',
        'boarding_station_sequence',
        'alighting_station_sequence',
        'adult_challenged_passenger_count',
        'adult_passenger_count',
        'child_challenged_passenger_count',
        'child_passenger_count',
    ]
    
    # Date fields
    DATE_FIELDS = [
        'ticket_valid_start_date',
        'ticket_valid_end_date',
    ]
    
    # DateTime fields
    DATETIME_FIELDS = [
        'boarding_at',
        'alighting_at',
        'payment_at',
    ]
    
    # Array fields (comma-separated in Excel)
    ARRAY_FIELDS = [
        'transfer_area_code_list',
        'transfer_station_code_list',
    ]

    def __init__(
        self, 
        scenario: Scenario, 
        ridership_upload: RidershipUpload, 
        validation_mode: str = None,
        max_tolerance_time: Optional[int] = None
    ):
        """
        Initialize the processor.
        
        Args:
            scenario: The scenario to process records for
            ridership_upload: The upload record
            validation_mode: Validation mode (railway, bus_ic, bus_cash)
            max_tolerance_time: Optional maximum time tolerance in minutes for trip matching.
                               If provided, search directly with this window (no expanding).
                               If None, use expanding window logic (5 -> 10 -> 15 -> 20).
        """
        self.scenario = scenario
        self.ridership_upload = ridership_upload
        self.success_records: List[RidershipRecord] = []
        self.error_records: List[RidershipUploadError] = []
        self.error_row_numbers: set = set()  # Track row numbers that have errors
        self.headers: List[str] = []
        self.header_to_index: Dict[str, int] = {}
        self.ridership_id_counter = 0
        self.max_tolerance_time = max_tolerance_time
        
        # Set validation mode and required fields
        if validation_mode and validation_mode in self.VALIDATION_MODE_CHOICES:
            self.validation_mode = validation_mode
        else:
            self.validation_mode = self.VALIDATION_MODE_RAILWAY  # Default
        
        self.required_fields = self.REQUIRED_FIELDS_BY_MODE.get(
            self.validation_mode, 
            self.REQUIRED_TIME_FIELDS
        )
        
        # Initialize GTFS data enricher for filling missing fields
        # Pass max_tolerance_time to the enricher
        self.gtfs_enricher = _GtfsDataEnricherImpl(scenario, max_tolerance_time=max_tolerance_time)

    def parse_excel(self, file) -> Tuple[int, int]:
        """
        Parse the Excel file and process all records
        Returns tuple of (success_count, error_count)
        """
        try:
            workbook = openpyxl.load_workbook(file, data_only=True)
            # Use the first sheet
            worksheet = workbook.active
            
            if worksheet is None:
                raise ValueError("Excelファイルにシートが見つかりません。")
            
            # Get headers from first row
            self._parse_headers(worksheet)
            
            if not self.headers:
                raise ValueError("Excelファイルにヘッダー行が見つかりません。")
            
            # Process data rows (starting from row 2)
            total_data_rows = 0
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                # Skip completely empty rows
                if all(cell is None or (isinstance(cell, str) and cell.strip() == '') for cell in row):
                    continue

                total_data_rows += 1
                self._process_row(row_idx, row)

            # Calculate success rows: total records saved minus those with errors
            total_saved_records = len(self.success_records)
            error_row_count = len(self.error_row_numbers)
            success_row_count = total_saved_records - error_row_count

            return success_row_count, error_row_count
            
        except InvalidFileException as e:
            logger.error(f"Invalid Excel file: {e}")
            raise ValueError(f"無効なExcelファイルです: {str(e)}")
        except Exception as e:
            logger.error(f"Error parsing Excel file: {e}")
            raise

    def parse_csv(self, file) -> Tuple[int, int]:
        """
        Parse the CSV file and process all records
        Returns tuple of (success_count, error_count)
        """
        import csv
        from io import StringIO
        
        try:
            # Handle different file types
            if hasattr(file, 'read'):
                # Check if it's a binary file or text file
                content = file.read()
                if isinstance(content, bytes):
                    # Try to decode with different encodings
                    text_content = None
                    for encoding in ['utf-8-sig', 'utf-8', 'shift_jis', 'cp932']:
                        try:
                            text_content = content.decode(encoding)
                            break
                        except UnicodeDecodeError:
                            continue
                    
                    if text_content is None:
                        raise ValueError("CSVファイルのエンコーディングを検出できません。UTF-8またはShift-JISを使用してください。")
                else:
                    text_content = content
            else:
                raise ValueError("無効なファイル形式です。")
            
            # Parse CSV content
            reader = csv.reader(StringIO(text_content))
            
            # Get headers from first row
            try:
                header_row = next(reader)
            except StopIteration:
                raise ValueError("CSVファイルが空です。")
            
            self._parse_csv_headers(header_row)
            
            if not self.headers:
                raise ValueError("CSVファイルにヘッダー行が見つかりません。")
            
            # Process data rows (starting from row 2)
            total_data_rows = 0
            for row_idx, row in enumerate(reader, start=2):
                # Skip completely empty rows
                if all(cell is None or (isinstance(cell, str) and cell.strip() == '') for cell in row):
                    continue

                total_data_rows += 1
                self._process_row(row_idx, tuple(row))

            # Calculate success rows: total records saved minus those with errors
            total_saved_records = len(self.success_records)
            error_row_count = len(self.error_row_numbers)
            success_row_count = total_saved_records - error_row_count

            return success_row_count, error_row_count
            
        except csv.Error as e:
            logger.error(f"CSV parsing error: {e}")
            raise ValueError(f"CSVファイルの解析中にエラーが発生しました: {str(e)}")
        except Exception as e:
            logger.error(f"Error parsing CSV file: {e}")
            raise

    def _parse_headers(self, worksheet):
        """Parse and validate headers from the first row (Excel)"""
        first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        
        self.headers = []
        self.header_to_index = {}
        
        for idx, cell in enumerate(first_row):
            if cell is not None:
                header = str(cell).strip().lower()
                self.headers.append(header)
                self.header_to_index[header] = idx
            else:
                self.headers.append(None)

    def _parse_csv_headers(self, header_row: list):
        """Parse and validate headers from the CSV header row"""
        self.headers = []
        self.header_to_index = {}
        
        for idx, cell in enumerate(header_row):
            if cell is not None and cell.strip():
                header = str(cell).strip().lower()
                self.headers.append(header)
                self.header_to_index[header] = idx
            else:
                self.headers.append(None)

    def _process_row(self, row_number: int, row: tuple):
        """Process a single row from the Excel/CSV file"""
        raw_data = self._row_to_dict(row)
        errors = []
        
        try:
            # Initialize record_data with empty strings for all CharField fields
            # This prevents NULL constraint violations for fields not in Excel
            record_data = self._get_default_record_data()
            
            for excel_col, model_field in self.COLUMN_MAPPING.items():
                if excel_col in self.header_to_index:
                    col_idx = self.header_to_index[excel_col]
                    if col_idx < len(row):
                        value = row[col_idx]
                        # Process the value based on field type
                        processed_value, field_errors = self._process_field_value(
                            model_field, value, row_number
                        )
                        # Only update if value is not None (preserve defaults)
                        if processed_value is not None or model_field in self.ARRAY_FIELDS:
                            record_data[model_field] = processed_value
                        errors.extend(field_errors)
            
            # ===== FALLBACK LOGIC: payment_at from alighting_at =====
            # payment_at can be same as alighting_at, but NOT boarding_at
            if self._is_field_empty(record_data.get('payment_at')):
                if not self._is_field_empty(record_data.get('alighting_at')):
                    # Use alighting_at as payment_at
                    record_data['payment_at'] = record_data['alighting_at']
            # ===== END FALLBACK LOGIC =====
            
            # Check required fields (time fields + conditional station fields)
            required_errors = self._validate_required_fields(record_data, row_number)
            errors.extend(required_errors)
            
            # Generate ridership_record_id if not provided
            if not record_data.get('ridership_record_id'):
                self.ridership_id_counter += 1
                record_data['ridership_record_id'] = self.ridership_id_counter
            
            # If there are validation errors, save to error table
            if errors:
                for error in errors:
                    self._add_error(
                        row_number=row_number,
                        error_type=error['type'],
                        field_name=error['field'],
                        error_message=error['message'],
                        raw_data=raw_data
                    )
            else:
                # Enrich record data with GTFS data for optional fields
                # This fills in trip_code, route_id, route_name, station names, and times
                # if they are not provided in the Excel
                boarding_station_code = record_data.get('boarding_station_code', '')
                alighting_station_code = record_data.get('alighting_station_code', '')
                boarding_station_name = record_data.get('boarding_station_name', '')
                alighting_station_name = record_data.get('alighting_station_name', '')
                payment_at = record_data.get('payment_at')
                boarding_at = record_data.get('boarding_at')
                alighting_at = record_data.get('alighting_at')

                record_data, trip_found = self.gtfs_enricher.enrich_record_data(
                    record_data=record_data,
                    boarding_station_code=boarding_station_code,
                    alighting_station_code=alighting_station_code,
                    payment_at=payment_at
                )

                # If trip matching failed, add error but still save the record
                if not trip_found:
                    # Format boarding/alighting time for error message
                    boarding_time_str = boarding_at.strftime('%Y-%m-%d %H:%M:%S') if boarding_at else '不明'
                    alighting_time_str = alighting_at.strftime('%Y-%m-%d %H:%M:%S') if alighting_at else '不明'

                    error_message = (
                        f"乗車駅 '{boarding_station_name or boarding_station_code}' と "
                        f"降車駅 '{alighting_station_name or alighting_station_code}' の間で、"
                        f"乗車時刻 {boarding_time_str} および降車時刻 {alighting_time_str} に "
                        f"一致するトリップが見つかりませんでした。"
                    )

                    self._add_error(
                        row_number=row_number,
                        error_type='trip_not_found',
                        field_name='trip_code',
                        error_message=error_message,
                        raw_data=raw_data
                    )

                # Create the record object (regardless of trip_found status)
                record = RidershipRecord(
                    scenario=self.scenario,
                    ridership_upload=self.ridership_upload,
                    source_row_number=row_number,
                    **record_data
                )
                self.success_records.append(record)
                
        except Exception as e:
            logger.error(f"Error processing row {row_number}: {e}")
            self._add_error(
                row_number=row_number,
                error_type='unknown',
                field_name='',
                error_message=f"行の処理中にエラーが発生しました: {str(e)}",
                raw_data=raw_data
            )
    
    def _get_default_record_data(self) -> Dict:
        """
        Initialize record data with default empty strings for all CharField fields.
        This prevents NULL constraint violations for fields with blank=True but not null=True.
        """
        return {
            # CharField fields - initialize with empty string
            'ic_card_agency_identification_code': '',
            'ic_card_issuer_code': '',
            'ic_card_issuer_name': '',
            'ic_card_feature_type': '',
            'ticket_type_area_code': '',
            'ticket_type': '',
            'ticket_type_name': '',
            'transportation_mode_code': '',
            'operating_agency_code': '',
            'operating_agency_name': '',
            'serviced_office_code': '',
            'serviced_office_name': '',
            'route_pattern_id': '',
            'route_pattern_number': '',
            'service_line_name': '',
            'route_id': '',
            'route_name': '',
            'trip_code': '',
            'timetable_number': '',
            'vehicle_number': '',
            'operation_type': '',
            'operation_detail_type': '',
            'boarding_area_code': '',
            'boarding_station_code': '',
            'boarding_station_name': '',
            'alighting_area_code': '',
            'alighting_station_code': '',
            'alighting_station_name': '',
            'passenger_classification_type': '',
            # IntegerField fields - initialize with None (will be processed by _process_integer)
            'ic_card_usage_detail_id': None,
            # ArrayField fields - initialize with empty list
            'transfer_area_code_list': [],
            'transfer_station_code_list': [],
        }

    def _process_field_value(
        self, field_name: str, value: Any, row_number: int
    ) -> Tuple[Any, List[Dict]]:
        """
        Process and validate a field value
        Returns (processed_value, list_of_errors)
        """
        errors = []
        
        # Handle None/empty values
        if value is None or (isinstance(value, str) and value.strip() == ''):
            # Return appropriate default for array fields
            if field_name in self.ARRAY_FIELDS:
                return [], errors
            return None, errors
        
        try:
            # Process based on field type
            if field_name in self.INTEGER_FIELDS:
                return self._process_integer(value, field_name, row_number, errors)
            
            elif field_name in self.DATE_FIELDS:
                return self._process_date(value, field_name, row_number, errors)
            
            elif field_name in self.DATETIME_FIELDS:
                return self._process_datetime(value, field_name, row_number, errors)
            
            elif field_name in self.ARRAY_FIELDS:
                return self._process_array(value, field_name, row_number, errors)
            
            elif field_name in self.ENUM_FIELDS:
                return self._process_enum(value, field_name, row_number, errors)
            
            else:
                # String field - convert to string
                return str(value).strip(), errors
                
        except Exception as e:
            errors.append({
                'type': 'parsing',
                'field': field_name,
                'message': f"フィールド '{field_name}' の値 '{value}' の処理中にエラーが発生しました: {str(e)}"
            })
            return None, errors

    def _process_integer(
        self, value: Any, field_name: str, row_number: int, errors: List[Dict]
    ) -> Tuple[Optional[int], List[Dict]]:
        """Process integer field value"""
        try:
            if isinstance(value, (int, float)):
                return int(value), errors
            elif isinstance(value, str):
                cleaned = value.strip()
                if cleaned:
                    return int(float(cleaned)), errors
            return None, errors
        except (ValueError, TypeError):
            errors.append({
                'type': 'invalid_format',
                'field': field_name,
                'message': f"フィールド '{field_name}' の値 '{value}' は整数に変換できません"
            })
            return None, errors

    def _process_date(
        self, value: Any, field_name: str, row_number: int, errors: List[Dict]
    ) -> Tuple[Optional[datetime], List[Dict]]:
        """Process date field value with flexible format support"""
        try:
            if isinstance(value, datetime):
                return value.date(), errors
            elif isinstance(value, date):
                return value, errors
            elif isinstance(value, str):
                cleaned = value.strip()
                if cleaned:
                    # Try with dateutil first (more flexible)
                    try:
                        dt = dateutil_parser.parse(cleaned, dayfirst=False, yearfirst=False)
                        return dt.date(), errors
                    except (ValueError, dateutil_parser.ParserError):
                        pass
                    
                    # Fallback to manual formats
                    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%m/%d/%Y']:
                        try:
                            return datetime.strptime(cleaned, fmt).date(), errors
                        except ValueError:
                            continue
                    
                    # If no format matches
                    errors.append({
                        'type': 'invalid_format',
                        'field': field_name,
                        'message': f"フィールド '{field_name}' の値 '{value}' は日付形式（YYYY-MM-DD）に変換できません"
                    })
            return None, errors
        except Exception as e:
            errors.append({
                'type': 'parsing',
                'field': field_name,
                'message': f"フィールド '{field_name}' の日付処理中にエラー: {str(e)}"
            })
            return None, errors

    def _process_datetime(
        self, value: Any, field_name: str, row_number: int, errors: List[Dict]
    ) -> Tuple[Optional[datetime], List[Dict]]:
        """
        Process datetime field value with flexible format support.

        All datetimes are interpreted and stored as JST (Asia/Tokyo).

        Supports formats including:
        - ISO format: 2025-04-01 06:00:00
        - Slash format: 2025/04/01 06:00
        - US format without leading zeros: 4/1/2025 6:00
        - Various other common formats via dateutil
        """
        try:
            if isinstance(value, datetime):
                # Make timezone aware with JST if not already
                if timezone.is_naive(value):
                    return value.replace(tzinfo=JST), errors
                return value, errors
            elif isinstance(value, str):
                cleaned = value.strip()
                if cleaned:
                    # Try parsing with dateutil.parser first (more flexible)
                    try:
                        # dayfirst=False for M/D/Y format (US format common in Excel)
                        # yearfirst=False to ensure year is not at front unless ISO format
                        dt = dateutil_parser.parse(cleaned, dayfirst=False, yearfirst=False)
                        if timezone.is_naive(dt):
                            dt = dt.replace(tzinfo=JST)
                        return dt, errors
                    except (ValueError, dateutil_parser.ParserError):
                        pass

                    # Fallback to manual formats if dateutil fails
                    formats = [
                        '%Y-%m-%d %H:%M:%S',   # 2025-04-01 06:00:00
                        '%Y-%m-%d %H:%M',       # 2025-04-01 06:00
                        '%Y/%m/%d %H:%M:%S',   # 2025/04/01 06:00:00
                        '%Y/%m/%d %H:%M',       # 2025/04/01 06:00
                        '%d/%m/%Y %H:%M:%S',   # 01/04/2025 06:00:00
                        '%d/%m/%Y %H:%M',       # 01/04/2025 06:00
                        '%m/%d/%Y %H:%M:%S',   # 04/01/2025 06:00:00
                        '%m/%d/%Y %H:%M',       # 04/01/2025 06:00
                        '%Y-%m-%dT%H:%M:%S',   # ISO format with T
                        '%Y-%m-%dT%H:%M',       # ISO format with T (no seconds)
                    ]
                    for fmt in formats:
                        try:
                            dt = datetime.strptime(cleaned, fmt)
                            return dt.replace(tzinfo=JST), errors
                        except ValueError:
                            continue

                    # If all formats fail
                    errors.append({
                        'type': 'invalid_format',
                        'field': field_name,
                        'message': f"フィールド '{field_name}' の値 '{value}' は日時形式に変換できません"
                    })
            return None, errors
        except Exception as e:
            errors.append({
                'type': 'parsing',
                'field': field_name,
                'message': f"フィールド '{field_name}' の日時処理中にエラー: {str(e)}"
            })
            return None, errors

    def _process_array(
        self, value: Any, field_name: str, row_number: int, errors: List[Dict]
    ) -> Tuple[List[str], List[Dict]]:
        """Process array field value (comma-separated string)"""
        try:
            if isinstance(value, list):
                return [str(v).strip() for v in value if v], errors
            elif isinstance(value, str):
                cleaned = value.strip()
                if cleaned:
                    # Split by comma, semicolon, or newline
                    items = []
                    for delimiter in [',', ';', '\n']:
                        if delimiter in cleaned:
                            items = [item.strip() for item in cleaned.split(delimiter) if item.strip()]
                            break
                    if not items and cleaned:
                        items = [cleaned]
                    return items, errors
            return [], errors
        except Exception as e:
            errors.append({
                'type': 'parsing',
                'field': field_name,
                'message': f"フィールド '{field_name}' の配列処理中にエラー: {str(e)}"
            })
            return [], errors

    def _process_enum(
        self, value: Any, field_name: str, row_number: int, errors: List[Dict]
    ) -> Tuple[Optional[str], List[Dict]]:
        """Process enum field value - store as-is without validation"""
        # Simply convert to string and strip whitespace
        # No enum validation - accept any value
        str_value = str(value).strip()
        return str_value, errors

    def _is_field_empty(self, value) -> bool:
        """Check if a field value is considered empty"""
        return (
            value is None or 
            (isinstance(value, str) and value.strip() == '') or
            (isinstance(value, list) and len(value) == 0)
        )

    def _validate_required_fields(
        self, record_data: Dict, row_number: int
    ) -> List[Dict]:
        """
        Validate that all required fields have values.
        
        Required fields:
        - boarding_at (time field - always required)
        - alighting_at (time field - always required)
        - boarding_station: code OR name must be present (conditional)
        - alighting_station: code OR name must be present (conditional)
        """
        errors = []
        
        # Validate time fields (strictly required)
        for field in self.required_fields:
            value = record_data.get(field)
            
            if self._is_field_empty(value):
                # Get Japanese display name if available
                display_name = self.FIELD_DISPLAY_NAMES.get(field, field)
                errors.append({
                    'type': 'missing_required',
                    'field': field,
                    'message': f"必須フィールド '{display_name}' ({field}) が設定されていません。"
                })
        
        # Validate conditional station fields (code OR name must be present)
        for code_field, name_field, display_name in self.CONDITIONAL_STATION_FIELDS:
            code_value = record_data.get(code_field)
            name_value = record_data.get(name_field)
            
            code_empty = self._is_field_empty(code_value)
            name_empty = self._is_field_empty(name_value)
            
            # Both code and name are empty - this is an error
            if code_empty and name_empty:
                code_display = self.FIELD_DISPLAY_NAMES.get(code_field, code_field)
                name_display = self.FIELD_DISPLAY_NAMES.get(name_field, name_field)
                errors.append({
                    'type': 'missing_required',
                    'field': f'{code_field}/{name_field}',
                    'message': f"'{display_name}'の情報が不足しています。'{code_display}' ({code_field}) または '{name_display}' ({name_field}) のいずれかを設定してください。"
                })
        
        return errors

    def _row_to_dict(self, row: tuple) -> Dict:
        """Convert a row tuple to a dictionary with header keys"""
        result = {}
        for idx, value in enumerate(row):
            if idx < len(self.headers) and self.headers[idx]:
                # Convert value to string representation for storage
                if value is not None:
                    if isinstance(value, datetime):
                        result[self.headers[idx]] = value.isoformat()
                    else:
                        result[self.headers[idx]] = str(value)
                else:
                    result[self.headers[idx]] = None
        return result

    def _add_error(
        self,
        row_number: int,
        error_type: str,
        field_name: str,
        error_message: str,
        raw_data: Dict
    ):
        """Add an error record and track the row number"""
        error = RidershipUploadError(
            ridership_upload=self.ridership_upload,
            source_row_number=row_number,
            error_type=error_type,
            field_name=field_name,
            error_message=error_message,
            raw_data=raw_data
        )
        self.error_records.append(error)
        self.error_row_numbers.add(row_number)  # Track row with error

    @transactional
    def save_records(self):
        """Save all success records and error records to database in batches"""
        batch_size = 1000  # Insert 1000 records at a time
        
        # Bulk create success records in batches
        if self.success_records:
            RidershipRecord.objects.bulk_create(
                self.success_records,
                batch_size=batch_size
            )
        
        # Bulk create error records in batches
        if self.error_records:
            RidershipUploadError.objects.bulk_create(
                self.error_records,
                batch_size=batch_size
            )


class RidershipRecordProcessor:
    """
    Stateless facade for ridership processing (SVC-003).

    Keeps the public API stateless while preserving the existing processing
    implementation for backward compatibility.
    """

    @staticmethod
    def parse_csv(
        *,
        scenario: Scenario,
        ridership_upload: RidershipUpload,
        file,
        validation_mode: str | None = None,
        max_tolerance_time: Optional[int] = None,
    ) -> tuple[_RidershipRecordProcessorImpl, int, int]:
        processor = _RidershipRecordProcessorImpl(
            scenario=scenario,
            ridership_upload=ridership_upload,
            validation_mode=validation_mode,
            max_tolerance_time=max_tolerance_time,
        )
        success_count, error_count = processor.parse_csv(file)
        return processor, success_count, error_count

    @staticmethod
    def parse_excel(
        *,
        scenario: Scenario,
        ridership_upload: RidershipUpload,
        file,
        validation_mode: str | None = None,
        max_tolerance_time: Optional[int] = None,
    ) -> tuple[_RidershipRecordProcessorImpl, int, int]:
        processor = _RidershipRecordProcessorImpl(
            scenario=scenario,
            ridership_upload=ridership_upload,
            validation_mode=validation_mode,
            max_tolerance_time=max_tolerance_time,
        )
        success_count, error_count = processor.parse_excel(file)
        return processor, success_count, error_count

    @staticmethod
    def save_records(*, processor: _RidershipRecordProcessorImpl) -> None:
        processor.save_records()

