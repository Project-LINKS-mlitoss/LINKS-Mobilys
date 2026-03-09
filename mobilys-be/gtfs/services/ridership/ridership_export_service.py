# Copyright (c) 2025-2026 MLIT Japan
# SPDX-License-Identifier: MIT
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Optional
from zoneinfo import ZoneInfo

import openpyxl
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# JST timezone for datetime display
JST = ZoneInfo("Asia/Tokyo")

from gtfs.constants import RIDERSHIP_EXPORT_COLUMNS as EXPORT_COLUMNS
from gtfs.models import RidershipRecord, RidershipUpload, RidershipUploadError, Scenario
from gtfs.services.base import log_service_call
from gtfs.services.ridership.errors import RidershipServiceError


@log_service_call
class RidershipExportService:
    @staticmethod
    def export(*, scenario_id: str, params: dict[str, Any]) -> HttpResponse:
        try:
            scenario = Scenario.objects.get(id=scenario_id)
        except Scenario.DoesNotExist as e:
            raise RidershipServiceError(
                message="シナリオが見つかりません",
                status_code=404,
                payload={"success": False, "message": "シナリオが見つかりません"},
            ) from e

        records = RidershipRecord.objects.filter(scenario=scenario)

        upload_id = params.get("upload_id", None)
        if upload_id:
            records = records.filter(ridership_upload_id=upload_id)

        start_date = params.get("start_date", None)
        end_date = params.get("end_date", None)

        if start_date:
            try:
                start_dt = datetime.strptime(start_date, "%Y-%m-%d")
                records = records.filter(payment_at__date__gte=start_dt.date())
            except ValueError:
                pass

        if end_date:
            try:
                end_dt = datetime.strptime(end_date, "%Y-%m-%d")
                records = records.filter(payment_at__date__lte=end_dt.date())
            except ValueError:
                pass

        boarding_station = params.get("boarding_station", None)
        if boarding_station:
            records = records.filter(boarding_station_code=boarding_station)

        alighting_station = params.get("alighting_station", None)
        if alighting_station:
            records = records.filter(alighting_station_code=alighting_station)

        include_empty_trips = params.get("include_empty_trips", "true").lower()
        if include_empty_trips == "false":
            records = records.exclude(trip_code="").exclude(trip_code__isnull=True)

        records = records.order_by("payment_at")

        record_count = records.count()
        if record_count == 0:
            raise RidershipServiceError(
                message="エクスポートするレコードがありません",
                status_code=404,
                payload={"success": False, "message": "エクスポートするレコードがありません"},
            )

        max_export_records = 100000
        if record_count > max_export_records:
            raise RidershipServiceError(
                message="エクスポート上限を超えています",
                status_code=400,
                payload={
                    "success": False,
                    "message": f"エクスポート上限を超えています。{record_count}件のレコードがありますが、上限は{max_export_records}件です。フィルターを使用してレコード数を減らしてください。",
                },
            )

        export_format = params.get("format", "xlsx").lower()
        if export_format == "csv":
            return RidershipExportService._export_csv(records, filename_prefix="ridership_export")
        return RidershipExportService._export_xlsx(records, filename_prefix="ridership_export")

    @staticmethod
    def export_by_upload(*, scenario_id: str, upload_id: str, params: dict[str, Any]) -> HttpResponse:
        try:
            scenario = Scenario.objects.get(id=scenario_id)
        except Scenario.DoesNotExist as e:
            raise RidershipServiceError(
                message="シナリオが見つかりません",
                status_code=404,
                payload={"success": False, "message": "シナリオが見つかりません"},
            ) from e

        try:
            upload = RidershipUpload.objects.get(id=upload_id, scenario=scenario)
        except RidershipUpload.DoesNotExist as e:
            raise RidershipServiceError(
                message="アップロードが見つかりません",
                status_code=404,
                payload={"success": False, "message": "アップロードが見つかりません"},
            ) from e

        records = RidershipRecord.objects.filter(scenario=scenario, ridership_upload=upload).order_by("source_row_number")
        if records.count() == 0:
            raise RidershipServiceError(
                message="エクスポートするレコードがありません",
                status_code=404,
                payload={"success": False, "message": "エクスポートするレコードがありません"},
            )

        export_format = params.get("format", "xlsx").lower()
        include_errors = params.get("include_errors", "false").lower() == "true"

        errors_qs = None
        if include_errors:
            errors_qs = RidershipUploadError.objects.filter(ridership_upload=upload).order_by("source_row_number")

        if export_format == "csv":
            return RidershipExportService._export_csv(records, filename_prefix="ridership", upload=upload)
        return RidershipExportService._export_xlsx(records, filename_prefix="ridership", upload=upload, errors=errors_qs)

    @staticmethod
    def _export_xlsx(
        records,
        *,
        filename_prefix: str,
        upload: Optional[RidershipUpload] = None,
        errors=None,
    ) -> HttpResponse:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "乗降実績データ"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, (_field_name, display_name) in enumerate(EXPORT_COLUMNS, start=1):
            cell = worksheet.cell(row=1, column=col_idx, value=display_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        data_alignment = Alignment(vertical="center")
        for row_idx, record in enumerate(records.iterator(), start=2):
            for col_idx, (field_name, _) in enumerate(EXPORT_COLUMNS, start=1):
                value = getattr(record, field_name, None)
                if value is not None:
                    if isinstance(value, datetime):
                        # Convert UTC to JST before formatting
                        if value.tzinfo is not None:
                            value = value.astimezone(JST)
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                    elif hasattr(value, "strftime"):
                        value = value.strftime("%Y-%m-%d")
                    elif isinstance(value, list):
                        value = ", ".join(str(v) for v in value) if value else ""

                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = data_alignment
                cell.border = thin_border

        for col_idx, (_field_name, display_name) in enumerate(EXPORT_COLUMNS, start=1):
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            width = max(12, min(30, len(display_name) * 2 + 2))
            worksheet.column_dimensions[column_letter].width = width

        worksheet.freeze_panes = "A2"

        if upload is not None:
            if errors is not None and errors.exists():
                ws_errors = workbook.create_sheet(title="エラー一覧")
                error_columns = [
                    ("source_row_number", "行番号"),
                    ("error_type", "エラー種別"),
                    ("field_name", "フィールド名"),
                    ("error_message", "エラーメッセージ"),
                ]

                for col_idx, (_field_name, display_name) in enumerate(error_columns, start=1):
                    cell = ws_errors.cell(row=1, column=col_idx, value=display_name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

                for row_idx, error in enumerate(errors.iterator(), start=2):
                    for col_idx, (field_name, _) in enumerate(error_columns, start=1):
                        value = getattr(error, field_name, None)
                        cell = ws_errors.cell(row=row_idx, column=col_idx, value=value)
                        cell.alignment = data_alignment
                        cell.border = thin_border

                ws_errors.column_dimensions["A"].width = 10
                ws_errors.column_dimensions["B"].width = 18
                ws_errors.column_dimensions["C"].width = 25
                ws_errors.column_dimensions["D"].width = 60
                ws_errors.freeze_panes = "A2"

            ws_summary = workbook.create_sheet(title="サマリー", index=0)
            summary_data = [
                ("アップロード名", upload.ridership_record_name),
                ("ファイル名", upload.file_name),
                ("アップロード日時", upload.uploaded_at.strftime("%Y-%m-%d %H:%M:%S")),
                ("処理完了日時", upload.processed_at.strftime("%Y-%m-%d %H:%M:%S") if upload.processed_at else ""),
                ("ステータス", upload.upload_status),
                ("総行数", upload.total_rows),
                ("成功行数", upload.success_rows),
                ("エラー数", upload.errors.count()),
                ("エクスポート日時", timezone.now().strftime("%Y-%m-%d %H:%M:%S")),
            ]

            summary_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            for row_idx, (label, value) in enumerate(summary_data, start=1):
                cell_label = ws_summary.cell(row=row_idx, column=1, value=label)
                cell_label.font = Font(bold=True)
                cell_label.fill = summary_fill
                cell_label.border = thin_border

                cell_value = ws_summary.cell(row=row_idx, column=2, value=value)
                cell_value.border = thin_border

            ws_summary.column_dimensions["A"].width = 20
            ws_summary.column_dimensions["B"].width = 40

        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        if upload is not None:
            safe_name = "".join(c for c in upload.ridership_record_name if c.isalnum() or c in ("_", "-", " "))[:30]
            filename = f"{filename_prefix}_{safe_name}_{timestamp}.xlsx"
        else:
            filename = f"{filename_prefix}_{timestamp}.xlsx"

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["X-Record-Count"] = str(records.count())
        return response

    @staticmethod
    def _export_csv(records, *, filename_prefix: str, upload: Optional[RidershipUpload] = None) -> HttpResponse:
        import csv
        from io import StringIO

        output = StringIO()
        writer = csv.writer(output)

        headers = [display_name for _, display_name in EXPORT_COLUMNS]
        writer.writerow(headers)

        for record in records.iterator():
            row = []
            for field_name, _ in EXPORT_COLUMNS:
                value = getattr(record, field_name, None)
                if value is not None:
                    if isinstance(value, datetime):
                        # Convert UTC to JST before formatting
                        if value.tzinfo is not None:
                            value = value.astimezone(JST)
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                    elif hasattr(value, "strftime"):
                        value = value.strftime("%Y-%m-%d")
                    elif isinstance(value, list):
                        value = ", ".join(str(v) for v in value) if value else ""
                else:
                    value = ""
                row.append(value)
            writer.writerow(row)

        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        if upload is not None:
            safe_name = "".join(c for c in upload.ridership_record_name if c.isalnum() or c in ("_", "-", " "))[:30]
            filename = f"{filename_prefix}_{safe_name}_{timestamp}.csv"
        else:
            filename = f"{filename_prefix}_{timestamp}.csv"

        response = HttpResponse("\ufeff" + output.getvalue(), content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["X-Record-Count"] = str(records.count())
        return response
